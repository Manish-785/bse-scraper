import os
import time
import json
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import (
    TimeoutException, WebDriverException, NoSuchElementException,
    StaleElementReferenceException, ElementNotInteractableException
)
from datetime import datetime
from dotenv import load_dotenv
import yfinance as yf
import requests
import re
import fitz
import openpyxl
import logging
import traceback
from functools import wraps
import sys
import psutil
import gc

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bse_monitor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# --- Configuration ---
MAX_RETRIES = 10
RETRY_DELAY = 30  # seconds
BATCH_SIZE = 10
BATCH_DELAY = 60  # seconds
MAIN_LOOP_DELAY = 60  # seconds
MEMORY_THRESHOLD = 1024 * 1024 * 1024  # 1GB in bytes
MAX_CONSECUTIVE_FAILURES = 10

# --- Error tracking ---
consecutive_failures = 0
last_success_time = datetime.now()

# --- Retry decorator ---
def retry_on_failure(max_retries=MAX_RETRIES, delay=RETRY_DELAY):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    logger.warning(f"{func.__name__} attempt {attempt + 1} failed: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))
                    else:
                        logger.error(f"{func.__name__} failed after {max_retries} attempts")
                        raise
            return None
        return wrapper
    return decorator

# --- Memory monitoring ---
def check_memory_usage():
    """Monitor memory usage and trigger cleanup if needed"""
    try:
        process = psutil.Process(os.getpid())
        memory_usage = process.memory_info().rss
        if memory_usage > MEMORY_THRESHOLD:
            logger.warning(f"High memory usage detected: {memory_usage / 1024 / 1024:.2f} MB")
            gc.collect()
            return True
    except Exception as e:
        logger.error(f"Memory check failed: {e}")
    return False

# --- Summarization Backends ---
from groq import Groq

# Environment variables with fallbacks
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

PROMPT_TEMPLATE = (
    "You are a financial analyst. Read the following BSE company update and write a one-line summary focusing only on the most important detail. "
    "Do not include any introductory or closing phrases, and do not write anything except the summary itself.\n\n{text}\n"
)

@retry_on_failure(max_retries=3, delay=5)
def summarise_with_groq_model(text, model, stream=False):
    """Summarize text using Groq API with error handling"""
    if not GROQ_API_KEY:
        raise ValueError("GROQ_API_KEY not available")
    
    try:
        client = Groq(api_key=GROQ_API_KEY)
        prompt = PROMPT_TEMPLATE.format(text=text)
        completion = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=256,
            top_p=1,
            stream=stream
        )
        
        if stream:
            result = ""
            for chunk in completion:
                part = chunk.choices[0].delta.content or ""
                result += part
            return result.strip()
        else:
            return completion.choices[0].message.content.strip()
    except Exception as e:
        logger.error(f"Groq API error with model {model}: {e}")
        raise

@retry_on_failure(max_retries=3, delay=5)
def summarise_with_gemini(text):
    """Summarize text using Gemini API with error handling"""
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY not available")
    
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
        headers = {"Content-Type": "application/json"}
        prompt = PROMPT_TEMPLATE.format(text=text)
        data = {"contents": [{"parts": [{"text": prompt}]}]}
        
        response = requests.post(url, headers=headers, json=data, timeout=30)
        if response.status_code == 429:
            raise Exception("Rate limit exceeded")
        response.raise_for_status()
        return response.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
    except Exception as e:
        logger.error(f"Gemini API error: {e}")
        raise

@retry_on_failure(max_retries=3, delay=10)
def summarise_with_openrouter(text):
    """Summarize text using OpenRouter API with error handling"""
    if not OPENROUTER_API_KEY:
        raise ValueError("OPENROUTER_API_KEY not available")
    
    try:
        from openai import OpenAI
        client = OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=OPENROUTER_API_KEY,
        )
        prompt = PROMPT_TEMPLATE.format(text=text)
        completion = client.chat.completions.create(
            model="deepseek/deepseek-r1-0528:free",
            messages=[{"role": "user", "content": prompt}],
            timeout=30
        )
        return completion.choices[0].message.content
    except Exception as e:
        logger.error(f"OpenRouter API error: {e}")
        raise

def summarise_bse_text(text, stream=False):
    """Try multiple summarization backends with fallbacks"""
    groq_models = [
        "llama3-70b-8192",
        "meta-llama/llama-4-scout-17b-16e-instruct",
        "llama-3.1-8b-instant",
    ]
    
    # Try Groq models first
    for model in groq_models:
        try:
            logger.info(f"Trying {model}...")
            summary = summarise_with_groq_model(text, model, stream=stream)
            if summary:
                logger.info(f"✔ Success with {model}")
                return summary
        except Exception as e:
            logger.warning(f"✘ Failed with {model}: {e}")
            continue

    # Fallback to Gemini
    try:
        logger.info("⚠ Falling back to Gemini...")
        summary = summarise_with_gemini(text)
        if summary:
            logger.info("✔ Success with Gemini")
            return summary
    except Exception as e:
        logger.warning(f"Gemini fallback failed: {e}")

    # Fallback to OpenRouter
    try:
        logger.info("⚠ Falling back to OpenRouter...")
        summary = summarise_with_openrouter(text)
        if summary:
            logger.info("✔ Success with OpenRouter")
            return summary
    except Exception as e:
        logger.warning(f"OpenRouter fallback failed: {e}")

    # Final fallback - return a basic summary
    logger.error("All summarization attempts failed, using fallback")
    return "Company announcement - details in PDF"

def chunk_text(text, max_chars=9000):
    """Split text into chunks of max_chars length, on paragraph boundaries if possible."""
    try:
        paragraphs = text.split('\n\n')
        chunks = []
        current = ""
        for para in paragraphs:
            if len(current) + len(para) + 2 <= max_chars:
                current += para + "\n\n"
            else:
                if current:
                    chunks.append(current.strip())
                current = para + "\n\n"
        if current:
            chunks.append(current.strip())
        return chunks
    except Exception as e:
        logger.error(f"Text chunking failed: {e}")
        return [text[:max_chars]]  # Return first part as fallback

@retry_on_failure(max_retries=3, delay=2)
def extract_for_summarization(pdf_path):
    """Extract text from PDF with error handling"""
    try:
        doc = fitz.open(pdf_path)
        formatted = ""
        for page in doc:
            try:
                blocks = page.get_text("dict")["blocks"]
                for block in blocks:
                    if "lines" in block:
                        for line in block["lines"]:
                            spans = [span["text"].strip() for span in line["spans"] if span["text"].strip()]
                            line_text = "  ".join(spans)
                            formatted += line_text + "\n"
                formatted += "\n"
            except Exception as e:
                logger.warning(f"Error processing page in PDF: {e}")
                continue
        doc.close()
        return formatted.strip() if formatted.strip() else "No text extracted"
    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
        return "PDF extraction failed"

def get_fo_stocks_from_excel(filename):
    """Get FO stocks from Excel file with error handling"""
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        fo_symbols = set()
        nse_col = None
        
        # Find the NSE Symbol column index
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and str(cell.value).strip().lower() == "nse symbol":
                nse_col = idx
                break
        
        if not nse_col:
            logger.warning("NSE Symbol column not found")
            return fo_symbols
        
        for row in ws.iter_rows(min_row=2):
            try:
                cell = row[nse_col - 1]
                # Check if cell has a fill color (not white or none)
                fill = cell.fill
                fgColor = fill.fgColor
                is_colored = False
                
                if fill and fgColor:
                    # Check for RGB color
                    if fgColor.type == 'rgb' and fgColor.rgb not in ('00000000', 'FFFFFFFF', 'FFFFFF', None):
                        is_colored = True
                    # Check for indexed color (not default)
                    elif fgColor.type == 'indexed' and fgColor.indexed not in (64, 0):
                        is_colored = True
                    # Check for theme color (not default)
                    elif fgColor.type == 'theme':
                        is_colored = True
                
                if is_colored and cell.value:
                    symbol = str(cell.value).strip()
                    if symbol:
                        fo_symbols.add(symbol)
            except Exception as e:
                logger.warning(f"Error processing row in Excel: {e}")
                continue
        
        wb.close()
        return fo_symbols
    except Exception as e:
        logger.error(f"Excel processing failed: {e}")
        return set()

@retry_on_failure(max_retries=3, delay=5)
def send_slack_notification(message, is_fo=False):
    """Send Slack notification with error handling"""
    SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL")
    if not SLACK_WEBHOOK_URL:
        logger.warning("SLACK_WEBHOOK_URL not set in environment.")
        return False
    
    try:
        if is_fo:
            payload = {
                "attachments": [
                    {
                        "color": "#36a64f",
                        "text": message,
                    },
                ],
                "username": "Event Bot",
                "icon_emoji": ":bell:"
            }
        else:
            payload = {
                "text": message,
                "username": "Event Bot",
                "icon_emoji": ":bell:"
            }
        
        response = requests.post(
            SLACK_WEBHOOK_URL,
            data=json.dumps(payload),
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        response.raise_for_status()
        return True
    except Exception as e:
        logger.error(f"Slack notification failed: {e}")
        raise

@retry_on_failure(max_retries=3, delay=5)
def download_pdf(url, filename):
    """Download PDF with error handling"""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        response = requests.get(url, headers=headers, allow_redirects=True, timeout=30)
        response.raise_for_status()
        
        with open(filename, 'wb') as f:
            f.write(response.content)
        return True
    except Exception as e:
        logger.error(f"PDF download failed for {url}: {e}")
        raise

def format_market_cap(market_cap):
    """Format market cap with error handling"""
    try:
        if market_cap is None or pd.isna(market_cap):
            return ""
        cr = market_cap / 1e7
        if cr >= 1e5:
            return f"₹{cr/1e5:.2f}L Cr"
        elif cr >= 1e3:
            return f"₹{cr/1e3:.2f}K Cr"
        else:
            return f"₹{cr:.2f} Cr"
    except Exception as e:
        logger.error(f"Market cap formatting failed: {e}")
        return ""

SENT_LINKS_FILE = "sent_links.txt"

def load_sent_links():
    """Load sent links with error handling"""
    try:
        if not os.path.exists(SENT_LINKS_FILE):
            return set()
        with open(SENT_LINKS_FILE, "r") as f:
            return set(line.strip() for line in f if line.strip())
    except Exception as e:
        logger.error(f"Loading sent links failed: {e}")
        return set()

def save_sent_link(link):
    """Save sent link with error handling"""
    try:
        with open(SENT_LINKS_FILE, "a") as f:
            f.write(link + "\n")
    except Exception as e:
        logger.error(f"Saving sent link failed: {e}")

def extract_timestamp_from_text(text):
    """Extract timestamp from text with error handling"""
    try:
        # Looks for DD-MM-YYYY HH:MM:SS
        match = re.search(r'(\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2})', text)
        if match:
            return match.group(1)
        return None
    except Exception as e:
        logger.error(f"Timestamp extraction failed: {e}")
        return None

def setup_driver():
    """Setup Chrome driver with error handling"""
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-plugins')
        options.add_argument('--disable-images')
        options.add_argument('--disable-javascript')
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        logger.error(f"Driver setup failed: {e}")
        raise

def safe_driver_quit(driver):
    """Safely quit driver"""
    try:
        if driver:
            driver.quit()
    except Exception as e:
        logger.warning(f"Driver quit failed: {e}")

def process_announcement(ann, market_cap, fo_symbols, sent_links):
    """Process a single announcement with comprehensive error handling"""
    try:
        name = ann.get('Name', '')
        link = ann.get('Link', '')
        
        # Skip checks
        skip_patterns = [
            "newspaper publication",
            "Compliances-Certificate under Reg. 74 (5) of SEBI (DP) Regulations, 2018",
            "materialisation", "dematerialisation", "materialized", "materialised",
            "dematerialized", "dematerialised", "materialization", "dematerialization"
        ]
        
        if any(pattern in name.lower() for pattern in skip_patterns):
            logger.info(f"Skipping announcement: {name}")
            return False
        
        if link in sent_links:
            logger.info(f"Skipping duplicate link: {link}")
            return False
        
        # Extract company code
        try:
            name_parts = name.split('-')
            code = name_parts[1].strip() if len(name_parts) > 1 else ""
        except Exception:
            code = ""
        
        # Get company info
        company_name = ""
        nse_symbol = None
        industry = None
        raw_market_cap = None
        
        try:
            mc_row = market_cap[market_cap['BSE Code'] == code]
            if not mc_row.empty:
                company_name = mc_row.iloc[0].get('Company Name', '')
                nse_symbol = mc_row.iloc[0].get('NSE Symbol', None)
                industry = mc_row.iloc[0].get('Industry', '')
                excel_market_cap = mc_row.iloc[0].get('Latest Market Cap', None)
                
                # Try to get market cap from Yahoo Finance
                if pd.notnull(nse_symbol):
                    try:
                        ticker_symbol = f"{nse_symbol}.NS"
                        ticker = yf.Ticker(ticker_symbol)
                        info = ticker.info
                        market_cap_yf = info.get('marketCap', None)
                        if market_cap_yf and market_cap_yf > 0:
                            raw_market_cap = market_cap_yf
                        elif excel_market_cap and excel_market_cap > 0:
                            raw_market_cap = excel_market_cap
                    except Exception as e:
                        logger.warning(f"Yahoo Finance data fetch failed for {ticker_symbol}: {e}")
                        if excel_market_cap and excel_market_cap > 0:
                            raw_market_cap = excel_market_cap
                elif excel_market_cap and excel_market_cap > 0:
                    raw_market_cap = excel_market_cap
        except Exception as e:
            logger.warning(f"Company info extraction failed: {e}")
        
        # Check market cap threshold
        try:
            if raw_market_cap is None or raw_market_cap < 500 * 1e7:
                logger.info(f"Skipping due to low market cap: {raw_market_cap} for {name}")
                return False
        except Exception:
            logger.info(f"Skipping due to invalid market cap for {name}")
            return False
        
        final_market_cap = format_market_cap(raw_market_cap)
        summary_name = company_name if company_name else name
        
        # Process PDF
        pdf_filename = f"temp_{int(time.time())}_{os.getpid()}.pdf"
        summary = "PDF processing failed"
        
        try:
            if download_pdf(link, pdf_filename):
                text = extract_for_summarization(pdf_filename)
                if text and text != "PDF extraction failed":
                    chunks = chunk_text(text, max_chars=9000)
                    if len(chunks) == 1:
                        summary = summarise_bse_text(chunks[0])
                    else:
                        logger.info(f"Text too long, splitting into {len(chunks)} chunks.")
                        chunk_summaries = []
                        for j, chunk in enumerate(chunks):
                            logger.info(f"Summarizing chunk {j+1}/{len(chunks)}...")
                            chunk_summary = summarise_bse_text(chunk)
                            chunk_summaries.append(chunk_summary)
                            time.sleep(2)  # Rate limiting
                        merged_summary_text = "\n".join(chunk_summaries)
                        summary = summarise_bse_text(merged_summary_text)
                else:
                    summary = "No text extracted from PDF"
        except Exception as e:
            logger.error(f"PDF processing failed for {name}: {e}")
            summary = "PDF processing failed"
        finally:
            # Clean up PDF file
            try:
                if os.path.exists(pdf_filename):
                    os.remove(pdf_filename)
            except Exception as e:
                logger.warning(f"PDF cleanup failed: {e}")
        
        # Send notification
        skip_patterns = [
            "newspaper publication",
            "Compliances-Certificate under Reg. 74 (5) of SEBI (DP) Regulations, 2018",
            "materialisation", "dematerialisation", "materialized", "materialised",
            "dematerialized", "dematerialised", "materialization", "dematerialization"
        ]
        
        if any(pattern in summarise_bse_text.lower() for pattern in skip_patterns):
            logger.info(f"Skipping announcement: {name}")
            return False
        try:
            release_time = ann.get('Time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            slack_message = (
                f"*{summary_name}* ({final_market_cap})\n"
                f"*Summary:* {summary}\n"
                f"*Industry:* {industry}\n"
                f"*Link:* {link}\n"
                f"*Release Time:* {release_time}\n"
            )
            
            is_fo = True if nse_symbol and str(nse_symbol).strip() in fo_symbols else False
            
            if send_slack_notification(slack_message, is_fo=is_fo):
                sent_links.add(link)
                save_sent_link(link)
                logger.info(f"Successfully processed and sent notification for: {summary_name}")
                
                # Send separator
                try:
                    send_slack_notification("-----\n\n")
                except Exception as e:
                    logger.warning(f"Separator notification failed: {e}")
                
                return True
            else:
                logger.error(f"Failed to send notification for: {summary_name}")
                return False
        except Exception as e:
            logger.error(f"Notification sending failed for {name}: {e}")
            return False
    
    except Exception as e:
        logger.error(f"Announcement processing failed: {e}")
        logger.error(traceback.format_exc())
        return False

def scrape_bse_announcements():
    """Scrape BSE announcements with comprehensive error handling"""
    driver = None
    try:
        driver = setup_driver()
        driver.get("https://www.bseindia.com/corporates/ann.html")
        wait = WebDriverWait(driver, 20)
        
        # Set dates
        from_date = to_date = datetime.today().strftime("%d/%m/%Y")
        
        # Wait for elements and set dates
        wait.until(EC.presence_of_element_located((By.ID, "txtFromDt")))
        wait.until(EC.presence_of_element_located((By.ID, "txtToDt")))
        
        driver.execute_script(f"document.getElementById('txtFromDt').value = '{from_date}';")
        driver.execute_script(f"document.getElementById('txtToDt').value = '{to_date}';")
        time.sleep(2)
        driver.execute_script("$('#txtFromDt').trigger('change');")
        driver.execute_script("$('#txtToDt').trigger('change');")
        
        # Select category
        category_select = wait.until(EC.presence_of_element_located((By.ID, "ddlPeriod")))
        select = Select(category_select)
        select.select_by_visible_text("Company Update")
        
        # Submit form
        submit_btn = wait.until(EC.element_to_be_clickable((By.ID, "btnSubmit")))
        driver.execute_script("arguments[0].scrollIntoView(true);", submit_btn)
        time.sleep(2)
        submit_btn.click()
        
        # Wait for results
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        time.sleep(3)
        
        new_announcements = []
        page_no = 1
        
        while True:
            try:
                # Extract announcements from current page
                for table in driver.find_elements(By.TAG_NAME, "table"):
                    for a_tag in table.find_elements(By.XPATH, ".//a[contains(@href, '.pdf')]"):
                        try:
                            href = a_tag.get_attribute('href')
                            if not href:
                                continue
                            
                            tr = a_tag.find_element(By.XPATH, "./ancestor::tr[1]")
                            first_td = tr.find_element(By.XPATH, "./td[1]")
                            name = first_td.text.strip()
                            
                            # Extract timestamp
                            timestamp_text = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
                            try:
                                parent = tr.find_element(By.XPATH, "..")
                                all_trs = parent.find_elements(By.XPATH, "./tr")
                                if len(all_trs) >= 2:
                                    second_last_tr = all_trs[-2]
                                    b_tags_in_second_last = second_last_tr.find_elements(By.TAG_NAME, "b")
                                    if b_tags_in_second_last:
                                        extracted_time = extract_timestamp_from_text(b_tags_in_second_last[0].text)
                                        if extracted_time:
                                            timestamp_text = extracted_time
                            except Exception as e:
                                logger.warning(f"Timestamp extraction failed: {e}")
                            
                            new_announcements.append({
                                'Name': name,
                                'Link': href,
                                'Time': timestamp_text
                            })
                        except Exception as e:
                            logger.warning(f"Error processing announcement link: {e}")
                            continue
                
                # Try to navigate to next page
                try:
                    next_btn = driver.find_element(By.ID, "idnext")
                    if next_btn.is_displayed() and next_btn.is_enabled():
                        driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
                        time.sleep(2)
                        next_btn.click()
                        page_no += 1
                        logger.info(f"Navigating to page {page_no}...")
                        time.sleep(3)
                        
                        # Wait for page to load
                        for _ in range(10):
                            try:
                                next_btn = driver.find_element(By.ID, "idnext")
                                if next_btn.is_enabled():
                                    break
                            except (StaleElementReferenceException, NoSuchElementException):
                                pass
                            time.sleep(1)
                    else:
                        logger.info("No more pages to navigate")
                        break
                except Exception as e:
                    logger.info(f"Navigation completed or failed: {e}")
                    break
            
            except Exception as e:
                logger.error(f"Error processing page {page_no}: {e}")
                break
        
        logger.info(f"Found {len(new_announcements)} announcements")
        return new_announcements
    
    except Exception as e:
        logger.error(f"BSE scraping failed: {e}")
        logger.error(traceback.format_exc())
        return []
    finally:
        safe_driver_quit(driver)

def main():
    """Main function with comprehensive error handling and recovery"""
    global consecutive_failures, last_success_time
    
    try:
        # Load environment variables
        load_dotenv()
        
        # Validate API keys
        if not os.getenv("OPENROUTER_API_KEY"):
            logger.error("OpenRouter API Key not found in .env file.")
            return
        
        # Load market cap data
        try:
            market_cap = pd.read_excel("Market Cap.xlsx")
            market_cap['BSE Code'] = (
                market_cap['BSE Code']
                .astype(str)
                .str.replace('.0', '', regex=False)
                .str.strip()
            )
            logger.info("Market cap data loaded successfully")
        except Exception as e:
            logger.error(f"Failed to load market cap data: {e}")
            return
        
        # Load FO symbols
        try:
            fo_symbols = get_fo_stocks_from_excel("Market Cap.xlsx")
            logger.info(f"Loaded {len(fo_symbols)} FO symbols")
        except Exception as e:
            logger.warning(f"Failed to load FO symbols: {e}")
            fo_symbols = set()
        
        # Load sent links
        sent_links = load_sent_links()
        logger.info(f"Loaded {len(sent_links)} sent links")
        
        logger.info("Starting BSE announcement monitoring...")
        
        while True:
            try:
                logger.info(f"[{datetime.now()}] Checking for new BSE announcements...")
                
                # Memory check
                check_memory_usage()
                
                # Scrape announcements
                new_announcements = scrape_bse_announcements()
                
                if new_announcements:
                    logger.info(f"Found {len(new_announcements)} new announcements")
                    
                    # Process announcements in batches
                    for i in range(0, len(new_announcements), BATCH_SIZE):
                        batch = new_announcements[i:i+BATCH_SIZE]
                        successful_processes = 0
                        
                        for ann in batch:
                            try:
                                if process_announcement(ann, market_cap, fo_symbols, sent_links):
                                    successful_processes += 1
                                time.sleep(2)  # Rate limiting between announcements
                            except Exception as e:
                                logger.error(f"Failed to process announcement: {e}")
                                continue
                        
                        logger.info(f"Batch {i//BATCH_SIZE + 1}: {successful_processes}/{len(batch)} successful")
                        
                        # Wait between batches (except for last batch)
                        if i + BATCH_SIZE < len(new_announcements):
                            logger.info(f"Waiting {BATCH_DELAY} seconds before next batch...")
                            time.sleep(BATCH_DELAY)
                    
                    # Reset failure counter on successful processing
                    consecutive_failures = 0
                    last_success_time = datetime.now()
                    
                else:
                    logger.info("No new announcements found")
                    consecutive_failures = 0  # Reset on successful scrape even if no announcements
                
            except Exception as e:
                consecutive_failures += 1
                logger.error(f"Main loop iteration failed (attempt {consecutive_failures}): {e}")
                logger.error(traceback.format_exc())
                
                # Check if we've had too many consecutive failures
                if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                    logger.critical(f"Too many consecutive failures ({consecutive_failures}). Last success: {last_success_time}")
                    
                    # Send emergency notification
                    try:
                        emergency_msg = (
                            f"🚨 *BSE Monitor Critical Error* 🚨\n"
                            f"Consecutive failures: {consecutive_failures}\n"
                            f"Last success: {last_success_time}\n"
                            f"Error: {str(e)[:200]}...\n"
                            f"Attempting to restart..."
                        )
                        send_slack_notification(emergency_msg)
                    except Exception as slack_e:
                        logger.error(f"Emergency notification failed: {slack_e}")
                    
                    # Wait longer before retry
                    time.sleep(RETRY_DELAY * 5)
                    consecutive_failures = 0  # Reset to try again
                else:
                    # Exponential backoff for failures
                    delay = min(RETRY_DELAY * (2 ** consecutive_failures), 300)  # Max 5 minutes
                    logger.info(f"Waiting {delay} seconds before retry...")
                    time.sleep(delay)
                    continue
            
            # Normal sleep between iterations
            logger.info(f"Sleeping for {MAIN_LOOP_DELAY} seconds before next check...")
            time.sleep(MAIN_LOOP_DELAY)
    
    except KeyboardInterrupt:
        logger.info("Received keyboard interrupt, shutting down gracefully...")
        sys.exit(0)
    
    except Exception as e:
        logger.critical(f"Critical error in main function: {e}")
        logger.critical(traceback.format_exc())
        
        # Send critical error notification
        try:
            critical_msg = (
                f"💀 *BSE Monitor Critical Failure* 💀\n"
                f"Script has stopped unexpectedly!\n"
                f"Error: {str(e)[:200]}...\n"
                f"Time: {datetime.now()}\n"
                f"Please check logs and restart manually."
            )
            send_slack_notification(critical_msg)
        except Exception as slack_e:
            logger.error(f"Critical notification failed: {slack_e}")
        
        # Try to restart after a delay
        logger.info("Attempting to restart in 60 seconds...")
        time.sleep(60)
        
        # Recursive restart (be careful with this)
        try:
            main()
        except Exception as restart_e:
            logger.critical(f"Restart failed: {restart_e}")
            sys.exit(1)

def health_check():
    """Perform health check and send status"""
    try:
        health_msg = (
            f"✅ *BSE Monitor Health Check* ✅\n"
            f"Status: Running\n"
            f"Time: {datetime.now()}\n"
            f"Last success: {last_success_time}\n"
            f"Consecutive failures: {consecutive_failures}\n"
            f"Memory usage: {psutil.Process().memory_info().rss / 1024 / 1024:.1f} MB"
        )
        send_slack_notification(health_msg)
        logger.info("Health check completed")
    except Exception as e:
        logger.error(f"Health check failed: {e}")

def startup_notification():
    """Send startup notification"""
    try:
        startup_msg = (
            f"🚀 *BSE Monitor Started* 🚀\n"
            f"Time: {datetime.now()}\n"
            f"Status: Monitoring BSE announcements\n"
            f"Batch size: {BATCH_SIZE}\n"
            f"Check interval: {MAIN_LOOP_DELAY}s"
        )
        send_slack_notification(startup_msg)
        logger.info("Startup notification sent")
    except Exception as e:
        logger.error(f"Startup notification failed: {e}")

if __name__ == "__main__":
    # Send startup notification
    startup_notification()
    
    # Set up periodic health checks (every 4 hours)
    import threading
    def periodic_health_check():
        while True:
            time.sleep(4 * 60 * 60)  # 4 hours
            health_check()
    
    # Start health check thread
    health_thread = threading.Thread(target=periodic_health_check, daemon=True)
    health_thread.start()
    
    # Run main function
    main()