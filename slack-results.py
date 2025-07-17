import io
import os
import time
import json
from typing import List, Optional, Tuple
import pandas as pd
import pdfplumber
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from dotenv import load_dotenv
import yfinance as yf
import requests
import fitz
import openpyxl
import re
import logging
import traceback
from logging.handlers import RotatingFileHandler
import signal
import sys
from pathlib import Path
import threading
from functools import wraps
import threading

def get_date_input(prompt_text, default, timeout=15):
    """
    Prompt user for date input with a timeout.
    If no input is received within 'timeout' seconds, return the default value.
    """
    user_input = [None]

    def ask():
        user_input[0] = input(f"{prompt_text} [{default}]: ").strip()

    thread = threading.Thread(target=ask)
    thread.daemon = True
    thread.start()
    thread.join(timeout)
    if thread.is_alive() or not user_input[0]:
        logger.info(f"No input received within {timeout} seconds. Using default: {default}")
        return default
    return user_input[0]

class FinancialReportExtractor:
    def __init__(self, pdf_link: str):
        self.pdf_links = [pdf_link]
        self.result_df = pd.DataFrame()
        # Enhanced search phrases with priority order
        self.SEARCH_PHRASES = [
            # Primary consolidated statements (most specific)
            "Statement of Audited Consolidated Financial Results",
            "Audited Consolidated Financial Results",
            "Statement of Unaudited Consolidated Financial Results", 
            "Unaudited Consolidated Financial Results",
            
            # Standalone financial statements
            "Statement of Audited Financial Results",
            "Audited Financial Results",
            "Statement of Unaudited Financial Results",
            "Unaudited Financial Results",
            
            # Consolidated statements variations
            "Consolidated Financial Results",
            "Statement of Consolidated Financial Results",
            "Consolidated Statement of Profit and Loss",
            "Consolidated Income Statement",
            "Consolidated Statement of Comprehensive Income",
            "Consolidated P&L Statement",
            
            # Standalone statements variations
            "Statement of Financial Results",
            "Statement of Profit and Loss",
            "Income Statement",
            "Statement of Comprehensive Income",
            "P&L Statement",
            "Profit and Loss Account",
            "Statement of Operations",
            "Statement of Earnings",
            
            # Quarterly/periodic results
            "Quarterly Results",
            "Quarterly Financial Results",
            "Quarterly Statement",
            "Half-yearly Results",
            "Half-yearly Financial Results",
            "Annual Results",
            "Annual Financial Results",
            "Interim Results",
            "Interim Financial Results",
            
            # Segment-wise results
            "Segment-wise Results",
            "Segmental Results",
            "Business Segment Results",
            "Geographic Segment Results",
            
            # Cash flow statements
            "Cash Flow Statement",
            "Statement of Cash Flows",
            "Consolidated Cash Flow Statement",
            "Statement of Cash Flow",
            
            # Balance sheet related
            "Balance Sheet",
            "Statement of Financial Position",
            "Consolidated Balance Sheet",
            "Statement of Assets and Liabilities",
            
            # Key financial metrics tables
            "Key Financial Ratios",
            "Financial Ratios",
            "Performance Indicators",
            "Key Performance Indicators",
            "Financial Highlights",
            "Financial Summary",
            "Operating Results",
            "Financial Performance",
            
            # Fallback phrases (broader)
            "Statement of Audited Consolidated",
            "Statement of Consolidated",
            "Audited Consolidated",
            "Unaudited Consolidated",
            "Consolidated Results",
            "Financial Results",
            "Results Summary",
            "Financial Statement",
            "Financial Information",
            "Financial Data"
        ]

        # Enhanced financial statement keywords to validate tables
        self.FINANCIAL_KEYWORDS = [
            # Revenue and income keywords
            'revenue', 'income', 'sales', 'turnover', 'net sales', 'gross sales',
            'operating revenue', 'total revenue', 'total income', 'other income',
            'interest income', 'dividend income', 'rental income',
            
            # Profit and loss keywords
            'profit', 'loss', 'net profit', 'gross profit', 'operating profit',
            'profit before tax', 'pbt', 'profit after tax', 'pat', 'net loss',
            'operating loss', 'comprehensive income', 'total comprehensive income',
            
            # Expense keywords
            'expenses', 'total expenses', 'operating expenses', 'cost of goods sold',
            'cogs', 'cost of sales', 'administrative expenses', 'selling expenses',
            'finance costs', 'financial expenses', 'interest expenses',
            'depreciation', 'amortization', 'impairment', 'provisions',
            
            # EBITDA and related metrics
            'ebitda', 'ebit', 'operating income', 'operating margin',
            'gross margin', 'net margin', 'profit margin',
            
            # Tax related
            'tax', 'income tax', 'current tax', 'deferred tax', 'tax expenses',
            'provision for tax', 'tax benefit',
            
            # Earnings metrics
            'earnings', 'earnings per share', 'eps', 'diluted eps', 'basic eps',
            'weighted average shares', 'dividend per share', 'dps',
            
            # Balance sheet items
            'assets', 'liabilities', 'equity', 'current assets', 'non-current assets',
            'total assets', 'current liabilities', 'non-current liabilities',
            'total liabilities', 'shareholders equity', 'retained earnings',
            'reserves', 'share capital', 'working capital',
            
            # Cash flow items
            'cash flow', 'operating cash flow', 'investing cash flow',
            'financing cash flow', 'free cash flow', 'cash and cash equivalents',
            'net cash', 'cash generated', 'cash used',
            
            # Financial ratios and metrics
            'return on equity', 'roe', 'return on assets', 'roa', 'debt to equity',
            'current ratio', 'quick ratio', 'debt ratio', 'asset turnover',
            'inventory turnover', 'receivables turnover',
            
            # Segment-wise keywords
            'segment', 'business segment', 'geographic segment', 'product segment',
            'division', 'subsidiary', 'joint venture', 'associate',
            
            # Time period indicators
            'quarter', 'quarterly', 'half-year', 'annual', 'year-to-date', 'ytd',
            'current year', 'previous year', 'corresponding period',
            
            # Common financial terms
            'total', 'subtotal', 'gross', 'net', 'operating', 'non-operating',
            'recurring', 'non-recurring', 'exceptional', 'extraordinary',
            'restated', 'standalone', 'consolidated', 'audited', 'unaudited',
            
            # Currency and units
            'crores', 'lakhs', 'millions', 'billions', 'thousands', 'rs', 'inr',
            'usd', 'eur', 'gbp', 'amount', 'value', 'figure'
        ]

        # Additional validation keywords for specific statement types
        self.STATEMENT_TYPE_KEYWORDS = {
            'profit_loss': [
                'profit', 'loss', 'income', 'revenue', 'expenses', 'ebitda',
                'depreciation', 'tax', 'earnings', 'margin'
            ],
            'balance_sheet': [
                'assets', 'liabilities', 'equity', 'capital', 'reserves',
                'current', 'non-current', 'total assets', 'total liabilities'
            ],
            'cash_flow': [
                'cash flow', 'operating activities', 'investing activities',
                'financing activities', 'cash generated', 'cash used'
            ],
            'ratios': [
                'ratio', 'percentage', 'times', 'days', 'turnover',
                'return', 'margin', 'coverage'
            ]
        }

        # Keywords to identify table headers and structure
        self.TABLE_STRUCTURE_KEYWORDS = [
            'particulars', 'description', 'items', 'details', 'account',
            'current year', 'previous year', 'march', 'september',
            'q1', 'q2', 'q3', 'q4', 'fy', 'financial year',
            'consolidated', 'standalone', 'audited', 'unaudited',
            'amount in', 'figures in', 'rs in', 'inr in'
        ]
            
        # Headers to avoid (these are usually not the main financial table)
        self.EXCLUDE_HEADERS = [
            'auditor', 'director', 'compliance', 'disclosure', 'notes',
            'segment', 'subsidiary', 'associate', 'investment'
        ]

    def download_pdf(self, url: str, timeout: int = 30) -> io.BytesIO:
        """Download PDF with better error handling and headers"""
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/pdf,*/*",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        }
        
        try:
            response = requests.get(url, headers=headers, timeout=timeout, stream=True)
            response.raise_for_status()
            
            # Check if it's actually a PDF
            content_type = response.headers.get('content-type', '').lower()
            if 'pdf' not in content_type and not url.lower().endswith('.pdf'):
                logger.warning(f"URL might not be a PDF: {url}")
            
            return io.BytesIO(response.content)
        except requests.exceptions.RequestException as e:
            logger.error(f"Error downloading {url}: {e}")
            raise
    
    def find_table_location(self, pdf_bytes: io.BytesIO) -> Optional[Tuple[int, float, str]]:
        """Find the best location for the primary consolidated financial table."""
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        best_match = None
        highest_score = 0

        try:
            for page_num in range(min(len(doc), 10)): # Check first 10 pages
                page = doc[page_num]
                page_text = page.get_text("text").lower()
                
                # Prioritize pages that contain both the title and header keywords
                has_title_phrase = any(phrase.lower() in page_text for phrase in self.SEARCH_PHRASES)
                has_financial_keywords = sum(1 for kw in ['particulars', 'revenue from operations', 'profit before tax', 'total income'] if kw in page_text) > 1

                if not has_title_phrase:
                    continue

                for phrase in self.SEARCH_PHRASES:
                    text_instances = page.search_for(phrase, quads=True)
                    if text_instances:
                        # Score based on phrase priority and presence of financial keywords
                        score = (len(self.SEARCH_PHRASES) - self.SEARCH_PHRASES.index(phrase)) * 2
                        score += has_financial_keywords * 5 # Give a high score if financial keywords are present

                        if score > highest_score:
                            highest_score = score
                            # Get the vertical position of the found phrase
                            y_position = max(q.rect.y1 for q in text_instances)
                            best_match = (page_num, y_position, phrase)
        finally:
            doc.close()

        if best_match:
            logger.info(f"Best table match found on page {best_match[0] + 1} with phrase: '{best_match[2]}'")
        else:
            logger.warning("Could not find a suitable table location.")
            
        return best_match
    
    def validate_table(self, table: List[List[str]], link: str) -> bool:
        """Validate if the extracted table is a legitimate financial statement."""
        if not table or len(table) < 4:  # Require at least a header and a few rows of data
            return False

        table_text = ' '.join([' '.join(cell for cell in row if cell) for row in table]).lower()
        
        # More stringent keyword check
        financial_keyword_count = sum(1 for keyword in self.FINANCIAL_KEYWORDS if keyword in table_text)
        
        # Check for presence of "Particulars" in the first column of the first few rows
        has_particulars = any("particulars" in str(row[0]).lower() for row in table[:3] if row)

        numeric_cells = sum(1 for row in table for cell in row if cell and self._is_numeric_value(cell))
        total_cells = sum(len(row) for row in table)
        numeric_ratio = numeric_cells / total_cells if total_cells > 0 else 0

        # Stricter validation criteria
        is_valid = (financial_keyword_count >= 3 or has_particulars) and numeric_ratio > 0.2 and len(table[0]) > 2
        logger.info(f"Table validation for {link}: Keywords found: {financial_keyword_count}, Numeric Ratio: {numeric_ratio:.2f}, Valid: {is_valid}")
        
        return is_valid
    
    def _calculate_phrase_score(self, phrase: str, page_text: str) -> float:
        """Calculate relevance score for a phrase match"""
        base_score = len(self.SEARCH_PHRASES) - self.SEARCH_PHRASES.index(phrase)
        
        # Bonus for financial keywords in context
        context_bonus = sum(1 for keyword in self.FINANCIAL_KEYWORDS 
                          if keyword in page_text) * 0.1
        
        # Penalty for exclusion keywords
        exclusion_penalty = sum(1 for keyword in self.EXCLUDE_HEADERS 
                              if keyword in page_text) * 0.2
        
        return base_score + context_bonus - exclusion_penalty

    def _is_numeric_value(self, value: str) -> bool:
        """Check if a value represents a number (including formatted numbers)"""
        if not value:
            return False
        
        # Remove common formatting
        clean_value = re.sub(r'[,\s()₹$]', '', str(value).strip())
        clean_value = re.sub(r'[^\d.-]', '', clean_value)
        
        try:
            float(clean_value)
            return True
        except ValueError:
            return False

    def extract_consolidated_table(self, pdf_bytes: io.BytesIO, max_pages: int = 5) -> Optional[List[List]]:
        """Extract consolidated financial table with improved logic"""
        
        # Find the best table location
        location = self.find_table_location(pdf_bytes)
        if not location:
            logger.warning("No matching phrases found")
            return None
            
        found_page, found_y, matched_phrase = location
        logger.info(f"Found table at page {found_page + 1}, phrase: '{matched_phrase}'")
        
        pdf_bytes.seek(0)
        tables = []
        header = None
        
        try:
            with pdfplumber.open(pdf_bytes) as pdf:
                # Extract from the found page
                page = pdf.pages[found_page]
                
                # Crop with some margin above the found text
                margin = 20  # pixels
                cropped = page.within_bbox((0, max(0, found_y - margin), 
                                          page.width, page.height))
                
                page_tables = cropped.extract_tables()
                if not page_tables:
                    logger.warning("No tables found in cropped area")
                    return None
                
                # Find the best table on the page
                best_table = self._select_best_table(page_tables)
                if not best_table:
                    logger.warning("No valid table found on the page")
                    return None
                
                tables.append(best_table)
                header = best_table[0] if best_table else None
                
                # Extract from subsequent pages if header matches
                for next_page_num in range(found_page + 1, 
                                         min(found_page + max_pages, len(pdf.pages))):
                    next_page = pdf.pages[next_page_num]
                    next_tables = next_page.extract_tables()
                    
                    if not next_tables:
                        break
                    
                    # Find matching table on next page
                    matching_table = self._find_matching_table(next_tables, header)
                    if matching_table:
                        tables.append(matching_table[1:])  # Skip header row
                        logger.info(f"Extended table from page {next_page_num + 1}")
                    else:
                        break  # Stop if no matching table found
                        
        except Exception as e:
            logger.error(f"Error extracting table: {e}")
            return None
        
        # Merge all table parts
        merged_table = self._merge_tables(tables)
        return merged_table if merged_table else None

    def _select_best_table(self, tables: List[List[List]]) -> Optional[List[List]]:
        """Select the most likely financial table from multiple tables"""
        if not tables:
            return None
            
        best_table = None
        best_score = 0
        
        for table in tables:
            if not table or len(table) < 2:
                continue
                
            # Score based on size and content
            size_score = min(len(table), 20) * 0.5  # Prefer reasonable size
            
            # Check for financial content
            table_text = ' '.join([' '.join(str(cell) for cell in row if cell) 
                                 for row in table[:5]]).lower()  # Check first 5 rows
            
            content_score = sum(1 for keyword in self.FINANCIAL_KEYWORDS 
                              if keyword in table_text)
            
            total_score = size_score + content_score
            
            if total_score > best_score:
                best_score = total_score
                best_table = table
                
        return best_table

    def _find_matching_table(self, tables: List[List[List]], 
                           target_header: List) -> Optional[List[List]]:
        """Find table with matching header structure"""
        if not target_header or not tables:
            return None
            
        for table in tables:
            if not table or len(table) < 1:
                continue
                
            table_header = table[0]
            
            # Check if headers match (allowing for some flexibility)
            if self._headers_match(table_header, target_header):
                return table
                
        return None

    def _headers_match(self, header1: List, header2: List, threshold: float = 0.7) -> bool:
        """Check if two headers match with some tolerance"""
        if not header1 or not header2:
            return False
            
        # Simple length check
        if abs(len(header1) - len(header2)) > 2:
            return False
            
        # Check content similarity
        h1_text = ' '.join(str(cell) for cell in header1 if cell).lower()
        h2_text = ' '.join(str(cell) for cell in header2 if cell).lower()
        
        if not h1_text or not h2_text:
            return False
            
        # Simple similarity check
        common_words = set(h1_text.split()) & set(h2_text.split())
        total_words = set(h1_text.split()) | set(h2_text.split())
        
        similarity = len(common_words) / len(total_words) if total_words else 0
        return similarity >= threshold

    def _merge_tables(self, tables: List[List[List]]) -> List[List]:
        """Merge multiple table parts into one"""
        if not tables:
            return []
            
        merged = []
        for i, table in enumerate(tables):
            if i == 0:
                merged.extend(table)
            else:
                merged.extend(table)  # Headers already removed for continuation pages
                
        return merged

    def clean_table_data(self, table: List[List]) -> List[List]:
        """Clean and standardize table data"""
        if not table:
            return table
            
        cleaned = []
        for row in table:
            cleaned_row = []
            for cell in row:
                if cell is None:
                    cleaned_row.append('')
                else:
                    # Clean cell content
                    cell_str = str(cell).strip()
                    # Remove excessive whitespace
                    cell_str = re.sub(r'\s+', ' ', cell_str)
                    cleaned_row.append(cell_str)
            cleaned.append(cleaned_row)
            
        return cleaned


# --- Enhanced Logging Setup ---
def setup_logging():
    """Setup comprehensive logging with rotation and multiple handlers"""
    # Create logs directory if it doesn't exist
    Path("logs").mkdir(exist_ok=True)
    
    # Configure root logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    
    # Clear existing handlers
    logger.handlers.clear()
    
    # Create formatters
    detailed_formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s'
    )
    simple_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # File handler for all logs (rotating)
    file_handler = RotatingFileHandler(
        'logs/bse_scraper.log',
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(detailed_formatter)
    
    # File handler for errors only
    error_handler = RotatingFileHandler(
        'logs/bse_scraper_errors.log',
        maxBytes=5*1024*1024,  # 5MB
        backupCount=3
    )
    error_handler.setLevel(logging.ERROR)
    error_handler.setFormatter(detailed_formatter)
    
    # Console handler for important messages
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(simple_formatter)
    
    # Add handlers to logger
    logger.addHandler(file_handler)
    logger.addHandler(error_handler)
    logger.addHandler(console_handler)
    
    return logger

# Initialize logger
logger = setup_logging()

# --- Health Check and Monitoring ---
class HealthMonitor:
    def __init__(self):
        self.last_successful_run = datetime.now()
        self.consecutive_failures = 0
        self.total_processed = 0
        self.total_errors = 0
        self.start_time = datetime.now()
        
    def record_success(self):
        self.last_successful_run = datetime.now()
        self.consecutive_failures = 0
        self.total_processed += 1
        
    def record_failure(self):
        self.consecutive_failures += 1
        self.total_errors += 1
        
    def get_health_status(self):
        uptime = datetime.now() - self.start_time
        time_since_success = datetime.now() - self.last_successful_run
        
        return {
            'uptime': str(uptime),
            'last_successful_run': self.last_successful_run.strftime('%Y-%m-%d %H:%M:%S'),
            'time_since_success': str(time_since_success),
            'consecutive_failures': self.consecutive_failures,
            'total_processed': self.total_processed,
            'total_errors': self.total_errors,
            'error_rate': f"{(self.total_errors / max(self.total_processed, 1)) * 100:.2f}%"
        }

health_monitor = HealthMonitor()

# --- Decorator for retry logic ---
def retry_on_failure(max_retries=3, delay=5, backoff=2):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    wait_time = delay * (backoff ** attempt)
                    logger.warning(f"{func.__name__} failed on attempt {attempt + 1}/{max_retries}: {e}")
                    if attempt < max_retries - 1:
                        logger.info(f"Retrying in {wait_time} seconds...")
                        time.sleep(wait_time)
                    else:
                        logger.error(f"{func.__name__} failed after {max_retries} attempts")
                        health_monitor.record_failure()
            raise last_exception
        return wrapper
    return decorator

# --- Graceful Shutdown Handler ---
class GracefulShutdown:
    def __init__(self):
        self.shutdown = False
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)
        
    def _signal_handler(self, signum, frame):
        logger.info(f"Received signal {signum}, initiating graceful shutdown...")
        self.shutdown = True
        
    def should_shutdown(self):
        return self.shutdown

shutdown_handler = GracefulShutdown()

# --- Enhanced Summarization with Better Error Handling ---
from groq import Groq

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

PROMPT_TEMPLATE = (
    "You are a financial analyst. Read the following BSE company update and write a one-line summary focusing only on the most important detail. "
    "Do not include any introductory or closing phrases, and do not write anything except the summary itself.\n\n{text}\n"
)

@retry_on_failure(max_retries=3, delay=3)
def summarise_with_groq_model(text, model, stream=False):
    if not GROQ_API_KEY:
        logger.error("GROQ_API_KEY not found")
        return None
        
    logger.debug(f"Attempting summarization with Groq model: {model}")
    client = Groq(api_key=GROQ_API_KEY)
    prompt = PROMPT_TEMPLATE.format(text=text)
    
    completion = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=256,
        top_p=1,
        stream=stream
    )
    
    if stream:
        result = ""
        for chunk in completion:
            part = chunk.choices[0].delta.content or ""
            result += part
        return result.strip()
    else:
        return completion.choices[0].message.content.strip()

@retry_on_failure(max_retries=3, delay=5)
def summarise_with_gemini(text):
    if not GEMINI_API_KEY:
        logger.error("GEMINI_API_KEY not found")
        return None
        
    logger.debug("Attempting summarization with Gemini")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
    headers = {"Content-Type": "application/json"}
    prompt = PROMPT_TEMPLATE.format(text=text)
    data = {"contents": [{"parts": [{"text": prompt}]}]}
    
    response = requests.post(url, headers=headers, json=data, timeout=30)
    
    if response.status_code == 429:
        logger.warning("Gemini rate limit hit")
        time.sleep(10)
        raise Exception("Rate limit exceeded")
        
    response.raise_for_status()
    return response.json()["candidates"][0]["content"]["parts"][0]["text"].strip()

@retry_on_failure(max_retries=5, delay=10)
def summarise_with_openrouter(text, api_key=OPENROUTER_API_KEY):
    if not api_key:
        logger.error("OPENROUTER_API_KEY not found")
        return None
        
    logger.debug("Attempting summarization with OpenRouter")
    from openai import OpenAI
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )
    prompt = PROMPT_TEMPLATE.format(text=text)
    
    completion = client.chat.completions.create(
        model="deepseek/deepseek-r1-0528:free",
        messages=[{"role": "user", "content": prompt}],
        timeout=30
    )
    return completion.choices[0].message.content

def summarise_bse_text(text, stream=False):
    """Enhanced summarization with comprehensive error handling"""
    logger.info("Starting text summarization")
    
    groq_models = [
        "llama3-70b-8192",
        "meta-llama/llama-4-scout-17b-16e-instruct",
        "llama-3.1-8b-instant",
    ]
    
    # Try Groq models
    for model in groq_models:
        try:
            logger.debug(f"Trying Groq model: {model}")
            summary = summarise_with_groq_model(text, model, stream=stream)
            if summary:
                logger.info(f"✔ Successfully summarized with {model}")
                return summary
        except Exception as e:
            logger.warning(f"✘ Failed with {model}: {e}")
            continue

    # Fallback to Gemini
    try:
        logger.debug("Falling back to Gemini")
        summary = summarise_with_gemini(text)
        if summary:
            logger.info("✔ Successfully summarized with Gemini")
            return summary
    except Exception as e:
        logger.warning(f"Gemini failed: {e}")

    # Fallback to OpenRouter
    try:
        logger.debug("Falling back to OpenRouter")
        summary = summarise_with_openrouter(text)
        if summary:
            logger.info("✔ Successfully summarized with OpenRouter")
            return summary
    except Exception as e:
        logger.warning(f"OpenRouter failed: {e}")

    logger.error("All summarization attempts failed")
    return "Summary generation failed - please check the source document"

def chunk_text(text, max_chars=9000):
    """Split text into chunks of max_chars length, on paragraph boundaries if possible."""
    logger.debug(f"Chunking text of length {len(text)}")
    paragraphs = text.split('\n\n')
    chunks = []
    current = ""
    for para in paragraphs:
        if len(current) + len(para) + 2 <= max_chars:
            current += para + "\n\n"
        else:
            if current:
                chunks.append(current.strip())
            current = para + "\n\n"
    if current:
        chunks.append(current.strip())
    logger.debug(f"Text split into {len(chunks)} chunks")
    return chunks

@retry_on_failure(max_retries=2, delay=2)
def extract_for_summarization(pdf_path):
    """Extract text from PDF with error handling"""
    logger.debug(f"Extracting text from PDF: {pdf_path}")
    
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    doc = fitz.open(pdf_path)
    formatted = ""
    
    try:
        for page_num, page in enumerate(doc):
            try:
                blocks = page.get_text("dict")["blocks"]
                for block in blocks:
                    if "lines" in block:
                        for line in block["lines"]:
                            spans = [span["text"].strip() for span in line["spans"] if span["text"].strip()]
                            line_text = "  ".join(spans)
                            formatted += line_text + "\n"
                formatted += "\n"
            except Exception as e:
                logger.warning(f"Error processing page {page_num}: {e}")
                continue
    finally:
        doc.close()
    
    result = formatted.strip() if formatted.strip() else None
    logger.debug(f"Extracted {len(result) if result else 0} characters from PDF")
    return result

def get_fo_stocks_from_excel(filename):
    """Load F&O stocks from Excel with error handling"""
    logger.debug(f"Loading F&O stocks from {filename}")
    
    if not os.path.exists(filename):
        logger.error(f"Excel file not found: {filename}")
        return set()
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        fo_symbols = set()
        nse_col = None
        
        # Find the NSE Symbol column index
        for idx, cell in enumerate(ws[1], 1):
            if str(cell.value).strip().lower() == "nse symbol":
                nse_col = idx
                break
        
        if not nse_col:
            logger.warning("NSE Symbol column not found in Excel file")
            return fo_symbols
        
        for row in ws.iter_rows(min_row=2):
            cell = row[nse_col - 1]
            # Check if cell has a fill color (not white or none)
            fill = cell.fill
            fgColor = fill.fgColor
            is_colored = False
            
            if fill and fgColor:
                # Check for RGB color
                if fgColor.type == 'rgb' and fgColor.rgb not in ('00000000', 'FFFFFFFF', 'FFFFFF', None):
                    is_colored = True
                # Check for indexed color (not default)
                elif fgColor.type == 'indexed' and fgColor.indexed not in (64, 0):
                    is_colored = True
                # Check for theme color (not default)
                elif fgColor.type == 'theme':
                    is_colored = True
            
            if is_colored:
                symbol = str(cell.value).strip()
                if symbol:
                    fo_symbols.add(symbol)
        
        logger.info(f"Loaded {len(fo_symbols)} F&O symbols")
        return fo_symbols
    
    except Exception as e:
        logger.error(f"Error loading F&O stocks: {e}")
        return set()

@retry_on_failure(max_retries=3, delay=2)
def extract_financial_table_and_save(pdf_link, result_df, df, temp_dir="temp_tables"):
    """
    Extracts the consolidated financial table from the PDF link, saves it as an Excel file,
    and returns the file path. Returns None if extraction fails.
    """
    try:
        os.makedirs(temp_dir, exist_ok=True)
        extractor = FinancialReportExtractor(pdf_link)
        extractor.result_df = result_df
        extractor.df = df
        
        pdf_bytes = extractor.download_pdf(pdf_link)
        raw_table = extractor.extract_consolidated_table(pdf_bytes)

        # The raw_table is a list of lists; it needs to be validated before cleaning.
        if raw_table and extractor.validate_table(raw_table):
            cleaned_df_table = extractor.clean_table_data(raw_table)
            
            if cleaned_df_table:
                if len(cleaned_df_table)> 0:
                    header = cleaned_df_table[0]
                    data_rows = cleaned_df_table[1:]
                    df_table = pd.DataFrame(data_rows, columns=header)
                else:
                    df_table = pd.DataFrame()
            
            if not df_table.empty:
                df_table['Source_Link'] = pdf_link
                try:
                    name_row = df[df['Link'] == pdf_link]['Name'].values[0]
                    bse_code = name_row.split('-')[1].strip()
                except Exception:
                    bse_code = ""

                if 'BSE Code' not in df_table.columns:
                    df_table.insert(0, "BSE Code", bse_code)
                else:
                    df_table['BSE Code'] = bse_code
                
                

                # Save the DataFrame to an Excel file
                excel_path = os.path.join(temp_dir, f"financial_table_{int(time.time())}.xlsx")
                df_table.to_excel(excel_path, index=False)
                logger.info(f"Successfully extracted and saved financial table to {excel_path}")
                return excel_path
            else:
                logger.warning(f"Table for {pdf_link} was empty after cleaning.")
                return None
        else:
            logger.info(f"No valid financial table found in {pdf_link}")
            return None
            
    except Exception as e:
        logger.error(f"Error extracting financial table for {pdf_link}: {e}", exc_info=True)
        return None

@retry_on_failure(max_retries=3, delay=2)
def send_slack_notification(message, is_fo=False):
    """Send Slack notification with retry logic"""
    SLACK_WEBHOOK_URL1 = os.getenv("SLACK_RESULT_WEBHOOK")
    
    if not SLACK_WEBHOOK_URL1:
        logger.error("SLACK_WEBHOOK_URL1 not set in environment")
        return False
    
    logger.debug(f"Sending Slack notification (F&O: {is_fo})")
    
    if is_fo:
        payload = {
            "attachments": [
                {
                    "color": "#36a64f",
                    "text": message,
                },
            ],
            "username": "Event Bot",
            "icon_emoji": ":bell:"
        }
    else:
        payload = {
            "text": message,
            "username": "Event Bot",
            "icon_emoji": ":bell:"
        }
    
    response = requests.post(
        SLACK_WEBHOOK_URL1,
        data=json.dumps(payload),
        headers={'Content-Type': 'application/json'},
        timeout=10
    )
    
    if response.status_code != 200:
        logger.error(f"Failed to send Slack notification: {response.status_code} - {response.text}")
        raise Exception(f"Slack notification failed: {response.status_code}")
    
    logger.debug("Slack notification sent successfully")
    return True

@retry_on_failure(max_retries=3, delay=2)
def send_slack_notification_with_file(message, file_path=None, is_fo=False):
    """
    Sends a Slack notification. If a file_path is provided, it uploads the file
    with the message as an initial comment. Otherwise, it sends a plain text message via webhook.
    """
    # --- Case 1: Uploading a file with a message using a Bot Token ---
    if file_path and os.path.exists(file_path):
        SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN")
        SLACK_CHANNEL = os.getenv("SLACK_CHANNEL") # Must be a Channel ID, not name

        if not all([SLACK_BOT_TOKEN, SLACK_CHANNEL]):
            logger.error("Cannot upload file: SLACK_BOT_TOKEN or SLACK_CHANNEL is not set in .env file.")
            logger.info("Falling back to sending a text-only notification.")
            return send_slack_notification(message, is_fo) # Fallback to webhook

        logger.info(f"Attempting to upload '{os.path.basename(file_path)}' to Slack channel {SLACK_CHANNEL}.")
        try:
            with open(file_path, "rb") as file_content:
                response = requests.post(
                    "https://slack.com/api/files.upload",
                    headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
                    data={
                        "channels": SLACK_CHANNEL,
                        "initial_comment": message,
                        "filename": os.path.basename(file_path)
                    },
                    files={"file": file_content},
                    timeout=30
                )
            response.raise_for_status()
            
            if not response.json().get("ok"):
                # This will log detailed errors from Slack like 'not_in_channel' or 'invalid_auth'
                logger.error(f"Failed to upload file to Slack. API Response: {response.text}")
                return False

            logger.info("File successfully uploaded to Slack with the summary message.")
            return True
        except Exception as e:
            logger.error(f"An exception occurred during file upload to Slack: {e}")
            return False

    # --- Case 2: Sending a text-only message using a Webhook ---
    else:
        if file_path:
            logger.warning(f"File path '{file_path}' was provided but file not found. Sending text-only notification.")
        return send_slack_notification(message, is_fo)

@retry_on_failure(max_retries=3, delay=2)
def download_pdf(url, filename):
    """Download PDF with retry logic"""
    logger.debug(f"Downloading PDF from {url}")
    
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers, allow_redirects=True, timeout=30)
    response.raise_for_status()
    
    with open(filename, 'wb') as f:
        f.write(response.content)
    
    logger.debug(f"PDF downloaded successfully: {filename}")

def format_market_cap(market_cap):
    """Format market cap with error handling"""
    if market_cap is None or pd.isna(market_cap):
        return ""
    
    try:
        cr = market_cap / 1e7
        if cr >= 1e5:
            return f"₹{cr/1e5:.2f}L Cr"
        elif cr >= 1e3:
            return f"₹{cr/1e3:.2f}K Cr"
        else:
            return f"₹{cr:.2f} Cr"
    except Exception as e:
        logger.warning(f"Error formatting market cap {market_cap}: {e}")
        return ""

SENT_LINKS_FILE = "sent_links.txt"

def load_sent_links():
    """Load sent links with error handling"""
    logger.debug("Loading sent links")
    
    if not os.path.exists(SENT_LINKS_FILE):
        logger.debug("Sent links file not found, creating set")
        return set()
    
    try:
        with open(SENT_LINKS_FILE, "r") as f:
            links = set(line.strip() for line in f if line.strip())
        logger.debug(f"Loaded {len(links)} sent links")
        return links
    except Exception as e:
        logger.error(f"Error loading sent links: {e}")
        return set()

def save_sent_link(link):
    """Save sent link with error handling"""
    try:
        with open(SENT_LINKS_FILE, "a") as f:
            f.write(link + "\n")
        logger.debug(f"Saved sent link: {link}")
    except Exception as e:
        logger.error(f"Error saving sent link: {e}")

def extract_timestamp_from_text(text):
    """Extract timestamp from text with error handling"""
    try:
        # Looks for DD-MM-YYYY HH:MM:SS
        match = re.search(r'(\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2})', text)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Error extracting timestamp: {e}")
    return None

def create_webdriver():
    """Create webdriver with comprehensive error handling"""
    logger.debug("Creating webdriver")
    
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-logging')
    options.add_argument('--log-level=3')
    
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        logger.debug("Webdriver created successfully")
        return driver
    except Exception as e:
        logger.error(f"Error creating webdriver: {e}")
        raise

def process_bse_page(driver, wait):
    """Process BSE page and extract announcements"""
    logger.info("Processing BSE page")
    
    # Set date filters
    today_date = datetime.today().strftime("%d/%m/%Y")
    from_date = get_date_input("Enter From Date (DD/MM/YYYY)", today_date)
    to_date = get_date_input("Enter To Date (DD/MM/YYYY)", today_date)
    
    wait.until(EC.presence_of_element_located((By.ID, "txtFromDt")))
    wait.until(EC.presence_of_element_located((By.ID, "txtToDt")))
    
    driver.execute_script(f"document.getElementById('txtFromDt').value = '{from_date}';")
    driver.execute_script(f"document.getElementById('txtToDt').value = '{to_date}';")
    time.sleep(1)
    
    driver.execute_script("$('#txtFromDt').trigger('change');")
    driver.execute_script("$('#txtToDt').trigger('change');")

    # Select category
    category_select = wait.until(EC.presence_of_element_located((By.ID, "ddlPeriod")))
    select = Select(category_select)
    select.select_by_visible_text("Result")

    # Submit form
    submit_btn = wait.until(EC.element_to_be_clickable((By.ID, "btnSubmit")))
    driver.execute_script("arguments[0].scrollIntoView(true);", submit_btn)
    time.sleep(1)
    submit_btn.click()

    wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
    time.sleep(2)

    new_announcements = []
    from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

    page_no = 1
    while True and not shutdown_handler.should_shutdown():
        logger.debug(f"Processing page {page_no}")
        
        for table in driver.find_elements(By.TAG_NAME, "table"):
            for a_tag in table.find_elements(By.XPATH, ".//a[contains(@href, '.pdf')]"):
                href = a_tag.get_attribute('href')
                if href:
                    try:
                        tr = a_tag.find_element(By.XPATH, "./ancestor::tr[1]")
                        first_td = tr.find_element(By.XPATH, "./td[1]")
                        name = first_td.text.strip()
                        
                        # Extract timestamp
                        timestamp_text = None
                        try:
                            parent = tr.find_element(By.XPATH, "..")
                            all_trs = parent.find_elements(By.XPATH, "./tr")
                            if len(all_trs) >= 2:
                                second_last_tr = all_trs[-2]
                                b_tags_in_second_last = second_last_tr.find_elements(By.TAG_NAME, "b")
                                if b_tags_in_second_last:
                                    timestamp_text = extract_timestamp_from_text(b_tags_in_second_last[0].text)
                        except Exception as e:
                            logger.debug(f"Error extracting timestamp: {e}")
                        
                        if not timestamp_text:
                            timestamp_text = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
                        
                        new_announcements.append({
                            'Name': name,
                            'Link': href,
                            'Time': timestamp_text
                        })
                        
                    except Exception as e:
                        logger.warning(f"Error processing announcement: {e}")
                        continue

        # Try to go to next page
        try:
            next_btn = driver.find_element(By.ID, "idnext")
            if next_btn.is_displayed() and next_btn.is_enabled():
                driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
                time.sleep(1)
                next_btn.click()
                page_no += 1
                time.sleep(2)
                
                # Wait for page to load
                for _ in range(10):
                    try:
                        next_btn = driver.find_element(By.ID, "idnext")
                        if next_btn.is_enabled():
                            break
                    except (StaleElementReferenceException, NoSuchElementException):
                        pass
                    time.sleep(1)
            else:
                logger.debug("No more pages to process")
                break
        except Exception as e:
            logger.debug(f"Error navigating to next page: {e}")
            break

    logger.info(f"Found {len(new_announcements)} total announcements")
    return new_announcements

def process_announcement(ann, market_cap, fo_symbols, sent_links,result_df=None, df=None):
    """Process individual announcement with comprehensive error handling"""
    name = ann['Name']
    link = ann['Link']
    
    if link in sent_links:
        logger.debug(f"Skipping duplicate link: {link}")
        return False
    
    logger.info(f"Processing announcement: {name}")
    
    try:
        # Extract company code
        try:
            name_parts = name.split('-')
            code = name_parts[1].strip()
        except Exception:
            code = ""
            logger.warning(f"Could not extract code from name: {name}")
        
        # Get company information
        company_name = ""
        final_market_cap = ""
        nse_symbol = None
        raw_market_cap = None
        industry = None
        
        if code:
            mc_row = market_cap[market_cap['BSE Code'] == code]
            if not mc_row.empty:
                company_name = mc_row.iloc[0]['Company Name']
                nse_symbol = mc_row.iloc[0]['NSE Symbol']
                industry = mc_row.iloc[0]['Industry']
                excel_market_cap = mc_row.iloc[0].get('Latest Market Cap', None)
                
                # Try to get market cap from Yahoo Finance
                market_cap_yf = None
                if pd.notnull(nse_symbol):
                    ticker_symbol = f"{nse_symbol}.NS"
                    try:
                        ticker = yf.Ticker(ticker_symbol)
                        info = ticker.info
                        market_cap_yf = info.get('marketCap', None)
                    except Exception as e:
                        logger.debug(f"Error fetching Yahoo Finance data for {ticker_symbol}: {e}")
                
                # Determine final market cap
                if market_cap_yf and market_cap_yf > 0:
                    raw_market_cap = market_cap_yf
                elif excel_market_cap and excel_market_cap > 0:
                    raw_market_cap = excel_market_cap
        
        # Check market cap threshold
        if raw_market_cap is None or raw_market_cap < 1000 * 1e7:
            logger.info(f"Skipping due to low market cap: {raw_market_cap} for {name}")
            return False
        
        final_market_cap = format_market_cap(raw_market_cap)
        summary_name = company_name if company_name else name
        
        # # Process PDF
        # pdf_filename = f"temp_{int(time.time())}_{os.getpid()}.pdf"
        # summary = "Error processing PDF"
        excel_path = None
        
        # try:
        #     download_pdf(link, pdf_filename)
        #     text = extract_for_summarization(pdf_filename)
            
            # if result_df is not None and df is not None:
            #     excel_path = extract_financial_table_and_save(link, result_df, df)
            
        #     if text:
        #         # Chunk if too large
        #         chunks = chunk_text(text, max_chars=9000)
        #         if len(chunks) == 1:
        #             summary = summarise_bse_text(chunks[0])
        #         else:
        #             logger.info(f"Text too long, splitting into {len(chunks)} chunks")
        #             chunk_summaries = []
        #             for j, chunk in enumerate(chunks):
        #                 logger.debug(f"Summarizing chunk {j+1}/{len(chunks)}")
        #                 chunk_summary = summarise_bse_text(chunk)
        #                 chunk_summaries.append(chunk_summary)
        #                 time.sleep(2)  # Rate limiting
                    
        #             merged_summary_text = "\n".join(chunk_summaries)
        #             logger.debug("Summarizing merged chunk summaries")
        #             summary = summarise_bse_text(merged_summary_text)
        #     else:
        #         summary = "No text extracted from PDF"
                
        # except Exception as e:
        #     logger.error(f"Error processing PDF for {name}: {e}")
        #     summary = f"Error processing document: {str(e)}"
        # finally:
        #     # Clean up PDF file
        #     try:
        #         if os.path.exists(pdf_filename):
        #             os.remove(pdf_filename)
        #     except Exception as e:
        #         logger.warning(f"Error removing PDF file: {e}")
        
        # Send notification
        release_time = ann.get('Time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        slack_message = (
            f"*{summary_name}* ({final_market_cap})\n"
            # f"*Summary:* {summary}\n"
            f"*Industry:* {industry}\n"
            f"*Link:* {link}\n"
            f"*Release Time:* {release_time}\n"
        )
        
        is_fo = nse_symbol and str(nse_symbol).strip() in fo_symbols
        
        if send_slack_notification_with_file(slack_message, file_path=excel_path, is_fo=is_fo):
            logger.info(f"excel file sent to Slack: {excel_path}")
            sent_links.add(link)
            save_sent_link(link)
            logger.info(f"Successfully processed and sent notification for: {summary_name}")
            health_monitor.record_success()
            # Clean up Excel file
            if excel_path and os.path.exists(excel_path):
                try:
                    os.remove(excel_path)
                except Exception as e:
                    logger.warning(f"Error removing Excel file: {e}")
            return True
        else:
            logger.error(f"Failed to send notification for: {summary_name}")
            return False

    except Exception as e:
        logger.error(f"Error processing announcement {name}: {e}")
        logger.error(traceback.format_exc())
        health_monitor.record_failure()
        return False

def log_health_status():
    """Log current health status"""
    status = health_monitor.get_health_status()
    logger.info(f"Health Status: {json.dumps(status, indent=2)}")

def main():
    """Main function with comprehensive error handling"""
    logger.info("Starting BSE Scraper with enhanced logging")
    
    # Load environment variables
    load_dotenv()
    
    # Validate required environment variables
    required_vars = ["OPENROUTER_API_KEY", "SLACK_RESULT_WEBHOOK"]
    missing_vars = [var for var in required_vars if not os.getenv(var)]
    
    if missing_vars:
        logger.error(f"Missing required environment variables: {missing_vars}")
        sys.exit(1)
    
    # Load market cap data
    try:
        market_cap = pd.read_excel("Market Cap.xlsx")
        market_cap['BSE Code'] = (
            market_cap['BSE Code']
            .astype(str)
            .str.replace('.0', '', regex=False)
            .str.strip()
        )
        logger.info(f"Loaded market cap data with {len(market_cap)} entries")
    except Exception as e:
        logger.error(f"Error loading market cap data: {e}")
        sys.exit(1)
    
    # Load F&O symbols
    fo_symbols = get_fo_stocks_from_excel("Market Cap.xlsx")
    
    # Load sent links
    sent_links = load_sent_links()
    
    # Start health monitoring thread
    def health_monitor_thread():
        while not shutdown_handler.should_shutdown():
            try:
                log_health_status()
                time.sleep(300)  # Log every 5 minutes
            except Exception as e:
                logger.error(f"Error in health monitor thread: {e}")
                time.sleep(60)
    
    health_thread = threading.Thread(target=health_monitor_thread, daemon=True)
    health_thread.start()
    
    # Main processing loop
    consecutive_failures = 0
    max_consecutive_failures = 5
    
    while not shutdown_handler.should_shutdown():
        try:
            logger.info(f"[{datetime.now()}] Starting new BSE check cycle")
            
            # Create webdriver
            driver = None
            try:
                driver = create_webdriver()
                driver.get("https://www.bseindia.com/corporates/ann.html")
                wait = WebDriverWait(driver, 20)
                
                # Process BSE page
                new_announcements = process_bse_page(driver, wait)
                
                # Filter out already sent links
                filtered_announcements = [
                    ann for ann in new_announcements 
                    if ann['Link'] not in sent_links
                ]
                
                logger.info(f"Found {len(filtered_announcements)} new announcements to process")
                
                # Process announcements in batches
                batch_size = 10
                for i in range(0, len(filtered_announcements), batch_size):
                    if shutdown_handler.should_shutdown():
                        break
                        
                    batch = filtered_announcements[i:i+batch_size]
                    logger.info(f"Processing batch {i//batch_size + 1} ({len(batch)} announcements)")
                    
                    for ann in batch:
                        if shutdown_handler.should_shutdown():
                            break
                            
                        try:
                            success = process_announcement(ann, market_cap, fo_symbols, sent_links,result_df=market_cap,df=market_cap)
                            if success:
                                # Send separator
                                send_slack_notification("-----\n\n")
                                time.sleep(2)  # Rate limiting
                        except Exception as e:
                            logger.error(f"Error processing individual announcement: {e}")
                            health_monitor.record_failure()
                            continue
                    
                    # Wait between batches (except for last batch)
                    # if i + batch_size < len(filtered_announcements) and not shutdown_handler.should_shutdown():
                    #     logger.info("Waiting 60 seconds before next batch...")
                    #     time.sleep(60)
                
                # Reset consecutive failures on successful cycle
                consecutive_failures = 0
                logger.info("BSE check cycle completed successfully")
                
            except Exception as e:
                logger.error(f"Error in BSE processing cycle: {e}")
                logger.error(traceback.format_exc())
                consecutive_failures += 1
                health_monitor.record_failure()
                
                if consecutive_failures >= max_consecutive_failures:
                    logger.critical(f"Too many consecutive failures ({consecutive_failures}). Implementing extended backoff.")
                    time.sleep(300)  # 5 minute extended backoff
                    consecutive_failures = 0
                
            finally:
                # Clean up webdriver
                if driver:
                    try:
                        driver.quit()
                        logger.debug("Webdriver closed successfully")
                    except Exception as e:
                        logger.warning(f"Error closing webdriver: {e}")
            
            # Wait before next cycle
            if not shutdown_handler.should_shutdown():
                sleep_time = 60 if consecutive_failures == 0 else min(60 * (2 ** consecutive_failures), 300)
                logger.info(f"Waiting {sleep_time} seconds before next cycle...")
                
                # Interruptible sleep
                for _ in range(sleep_time):
                    if shutdown_handler.should_shutdown():
                        break
                    time.sleep(1)
                
        except KeyboardInterrupt:
            logger.info("Received keyboard interrupt, shutting down...")
            break
        except Exception as e:
            logger.error(f"Unexpected error in main loop: {e}")
            logger.error(traceback.format_exc())
            consecutive_failures += 1
            health_monitor.record_failure()
            
            # Emergency sleep
            time.sleep(30)
    
    logger.info("BSE Scraper shutting down gracefully")
    log_health_status()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.critical(f"Critical error in main: {e}")
        logger.critical(traceback.format_exc())
        sys.exit(1)